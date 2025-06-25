from openpyxl import load_workbook

# 读取Excel文件
workbook = load_workbook('e:\\github_code_lib\\gaokao_school\\souce\\school_mapping.xlsx')
sheet = workbook.active

print('工作表名称:', workbook.sheetnames)
print('数据行数:', sheet.max_row)
print('数据列数:', sheet.max_column)

# 读取前几行数据
print('\n前10行数据:')
for row in range(1, min(11, sheet.max_row + 1)):
    row_data = []
    for col in range(1, sheet.max_column + 1):
        cell_value = sheet.cell(row=row, column=col).value
        row_data.append(cell_value)
    print(f'第{row}行:', row_data)

# 如果数据不多，显示所有数据
if sheet.max_row <= 50:
    print('\n所有数据:')
    for row in range(1, sheet.max_row + 1):
        row_data = []
        for col in range(1, sheet.max_column + 1):
            cell_value = sheet.cell(row=row, column=col).value
            row_data.append(cell_value)
        print(f'第{row}行:', row_data)